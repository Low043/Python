from lowModel.utils import *
from lowModel.menus import *
import openpyxl as xl
import win32com.client
from math import ceil
from shutil import copyfile
from typing import Literal

from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from pypdf import PdfMerger
from pypdf import PdfReader
from pdf2image import convert_from_path
import pytesseract
import glob

class Excel:
    def __init__(self,pathToXlsx:str):
        self.wb = xl.load_workbook(pathToXlsx)#OpenPyXl WorkBook
        self.path = pathToXlsx
        self.fileName = self.__getFileNameByFilePath(pathToXlsx)
        self.folderPath = self.__getFolderPathByFilePath(pathToXlsx)
        self.savesBackupFolder = self.__getBackupFolderByPath(pathToXlsx)

    def __getFileNameByFilePath(self,path:str):#Returns the text between the last slash and the .xlsx (FileName)
        return path[path.rfind('\\')+1:path.rfind('.')]
        
    def __getFolderPathByFilePath(self,path:str):#Returns all text before the last slash (FolderPath)
        return path[:path.rfind('\\')+1]
    
    def __getBackupFolderByPath(self,path:str):#Returns the path to a new folder where backups of the modified Excel will be stored
        return self.__getFolderPathByFilePath(path) + '0oldVersions\\'

    def getSheets(self) -> list:#Returns the name of WorkSheets
        return self.wb.sheetnames
    
    def getRow(self,sheet:str,row:int):#Returns a list of Cells in Row (Begins with 1)
        return list(self.wb[sheet].rows)[row-1]
    
    def getColumn(self,sheet:str,column):#Returns a list of Cells in Column (Begins with 1)
        column = self.convertColumn(column)
        return list(self.wb[sheet].columns)[column-1]
    
    def getRowOfValue(self,sheet:str,column,value,occurrenceIndex=1):#Return the Row of Value in Column (Index -1 returns all occurrences)
        column = self.convertColumn(column)
        occurrences = []
        for row,cell in enumerate(self.getColumn(sheet,column),1):
            if cell.value == value:
                occurrences.append(row)
                if len(occurrences) == occurrenceIndex:
                    return row
        if occurrenceIndex == -1:
            return occurrences
        return None
    
    def getColumnOfValue(self,sheet:str,row:int,value,occurrenceIndex=1):#Return the Column of Value in Row (Index -1 returns all occurrences)
        occurrences = []
        for column,cell in enumerate(self.getRow(sheet,row),1):
            if cell.value == value:
                occurrences.append(column)
                if len(occurrences) == occurrenceIndex:
                    return column
        if occurrenceIndex == -1:
            return occurrences
        return None
    
    def getCellValue(self,sheet:str,column,row:int,allowFormula=True):#Returns Cell Value (Row begins with 1 and Column can be int or str)
        column = self.convertColumn(column)
        if allowFormula:
            return self.wb[sheet].cell(row,column).value
        else:
            readWb = xl.load_workbook(self.path,read_only=True,data_only=True)
            result = readWb[sheet].cell(row,column).value
            readWb.close()
            return result
        
    def setCellValue(self,sheet:str,column,row:int,value):#Change Cell Value (needs "save()")
        column = self.convertColumn(column)
        self.wb[sheet].cell(row,column).value = value

    def getCellFontStyle(self,sheet:str,column,row:int) -> Font:#Get Font Style from Cell
        column = self.convertColumn(column)
        return self.wb[sheet].cell(row,column).font

    def setCellFontStyle(self,sheet:str,column,row:int,fontStyle:Font):#Set a cell Font Style
        column = self.convertColumn(column)#Cell.Font doesn't allow fontStyle but allow a new Font equivalent
        self.wb[sheet].cell(row,column).font = Font(fontStyle.name,fontStyle.sz,fontStyle.b,fontStyle.i,fontStyle.charset,fontStyle.u,fontStyle.strike,fontStyle.color,fontStyle.scheme,fontStyle.family,fontStyle.size)

    #MAKE AUTO SIZE IF -1 !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    def setSize(self,sheet:str,column=None,row=None,size=-1):#Set a row/column height/width
        if row:
            self.wb[sheet].row_dimensions[row].height = size
        if column:
            column = self.convertColumn(column,toStr=True)
            self.wb[sheet].column_dimensions[column].width = size

    def setHide(self,sheet:str,column=None,row=None,hide=True):#Hide or Unhide a row/column
        if row:
            self.wb[sheet].row_dimensions[row].hidden = hide
        if column:
            column = self.convertColumn(column,toStr=True)
            self.wb[sheet].column_dimensions[column].hidden = hide

    def save(self,path=None,backup=True,pagesPdf=[-1]):#Save file changes in format of path (if path==None replace original file)
        if path == None:
            path = self.path
        if backup:#If backup==True create a copy of file in 'filePath/0oldVersions/'
            self.__backup()

        if path[-4:] == '.pdf':#If path ends with .pdf save file as PDF
            self.wb.save(self.path)
            self.__savePdf(path,pagesPdf)
        elif path[-5:] == '.xlsx':#If path ends with .xlsx save file as Excel
            self.wb.save(path)#OpenPyXl function to save Excel

    def __backup(self):#Create a copy of Excel in Backup folder
        if not os.path.exists(self.savesBackupFolder):#If backup folder doesn't exist, create it
            os.mkdir(self.savesBackupFolder)

        backupFilePath = self.savesBackupFolder + self.fileName + '.xlsx'

        version = 0
        while os.path.exists(backupFilePath):#If always exists a file with this name in this folder change the file name
            version += 1
            backupFilePath = self.savesBackupFolder + self.fileName + f' ({version})' + '.xlsx'

        copyfile(self.path,backupFilePath)#Copy file

    def __savePdf(self,path:str,pagesPdf:list):#Save file as .pdf
        if path == None:
            path = self.path.replace('.xlsx','.pdf')
        
        #The code below was copied from StackOverflow so i don't know how it works
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(self.path)
        wb.WorkSheets(self.__convertPages(pagesPdf)).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0,path)
        os.system("taskkill /f /im excel.exe")#Kill Excel after save

    def __convertPages(self,pages:list):#Replace page name by page index
        if pages == [-1]:#If pages==[-1] (default value), returns a list [1...n] where n is the index of last page in file
            return range(1,len(self.getSheets())+1)
        
        for n,page in enumerate(pages):
            if type(page) == str and page in self.getSheets():#If page is the name of some Sheet, change value to sheet index
                pages[n] = self.getSheets().index(page)
        return pages

    def convertColumn(self,column,toStr=False):#Converts column to Str or Int
        if type(column) == str and toStr == False:
            return ord(column.upper())+1 - ord('A')
        if type(column) == int and toStr == True:
            return chr(column-1+ord('A'))
        return column
    
    def close(self):#Close Excel file REMEMBER IT!!!!
        self.wb.close()
    
    def getCellRange(self,sheet:str,column1,row1:int,column2,row2:int,allowFormula=True):#Returns a list of Cell Values in range
        column1, column2 = self.convertColumn(column1), self.convertColumn(column2)
        values = []
        for column in range(column1,column2+1):
            values.append([])
            for row in range(row1,row2+1):
                values[-1].append(self.getCellValue(sheet,column,row,allowFormula))
        return values
    
    def setCellRange(self,sheet:str,column1,row1:int,column2,row2:int,rangeValues:list):#Set Cell Values in range based in RangeValues
        column1, column2 = self.convertColumn(column1), self.convertColumn(column2)
        for column in range(column1,column2+1):
            for row in range(row1,row2+1):
                self.setCellValue(sheet,column,row,rangeValues[column-column1][row-row1])

    class Table:#Object to manipulate cells inside a Table
        def __init__(self,excel,sheet:str,headerRow:int,mainColumn=1,lastRow=-1):
            self.excel = excel
            self.sheet = sheet
            self.headerRow = headerRow
            self.mainColumn = mainColumn - 1
            self.columns = self.__getColumns()
            self.firstRow = self.__getRow('First')
            self.lastRow = self.__getRow('Last') if lastRow == -1 else lastRow
            self.linkMethods = {
                'Name' : self.__linkByName,
                'NameSimilarity' : lambda x,y: self.__linkByName(x,y,similar=True),
                'Order' : self.__linkByOrder}

        class Column:
            def __init__(self,name:str,index:int):
                self.name = name
                self.index = index

        def __getColumns(self) -> list[Column]:
            columns = []
            for column,headerCell in enumerate(self.excel.getRow(self.sheet,self.headerRow),start=1):
                if headerCell.value != None:
                    columns.append(self.Column(headerCell.value,self.excel.convertColumn(column,toStr=True)))
            return columns
        
        def __getRow(self,row:str):#Returns the row of first value (returns the row of last value if recieve "Last")
            mainColumn = self.excel.getColumn(self.sheet,self.columns[self.mainColumn].index)
            mainColumn = reversed(mainColumn) if row == 'Last' else mainColumn
            
            for cell in mainColumn:
                if cell.row > self.headerRow and cell.value != None:
                    return cell.row
            return self.headerRow + 1

        _str = Literal['Name','NameSimilarity','Order']#Default values suggestion to linkBy
        def pullValuesFrom(self,table,replaceOldValues=False,linkBy:list[_str]=['Name'],cut=False,associate:dict={},notAssociate:dict={}):
            linkedColumns = {}#List of linked columns
            self.__columnsRemain = self.columns[:]#List of columns that not linked yet
            self.__pullColumnsRemain = table.columns[:]

            associate = self.__convertAssociations(table,associate)
            notAssociate = self.__convertAssociations(table,notAssociate)
            
            for association in associate:#Link manual associated columns and remove it from remain columns
                linkedColumns[association] = associate[association]
                self.__columnsRemain.remove(association)
                self.__pullColumnsRemain.remove(associate[association])

            for method in linkBy:#Use linkBy methods in list
                linkedColumns = self.linkMethods[method](linkedColumns,notAssociate)

            self.__pullColumns(table,linkedColumns,replaceOldValues,cut)

        def updateValuesOf(self,column,usingFunction=None,usingColumn=-1):
            column = self.__matchColumn(column)
            usingColumn = self.__matchColumn(usingColumn) if usingColumn != -1 else column
            usingFunction = (lambda cellValue:cellValue) if usingFunction == None else usingFunction

            if column != None and usingColumn != None:
                for cell in self.excel.getColumn(self.sheet,column.index):
                    if cell.row >= self.firstRow and cell.row <= self.lastRow:
                        cell.value = usingFunction(self.excel.getCellValue(self.sheet,usingColumn.index,cell.row))

        def validateValuesOf(self,column,usingTable,autoCorrectWhen=0.9):
            column = self.__matchColumn(column)
            allowedValues = usingTable.__getValuesFromMainColumn()
            for cell in self.excel.getColumn(self.sheet,column.index):
                if cell.value != None and cell.row >= self.firstRow and cell.row <= self.lastRow:
                    cell.value = self.__getMoreSimilar(cell.value,allowedValues,autoCorrectWhen)

        def __linkByName(self,linkedColumns={},notAssociate={},similar=False):
            for selfColumn in self.__columnsRemain:
                for tableColumn in self.__pullColumnsRemain:
                    namesMatch = (selfColumn.name == tableColumn.name) if similar == False else (textSimilarity(str(selfColumn.name),str(tableColumn.name)) >= 0.8)
                    if namesMatch and not (selfColumn in notAssociate and notAssociate[selfColumn] == tableColumn):
                        linkedColumns[selfColumn] = tableColumn
                        self.__columnsRemain.remove(selfColumn)
                        self.__pullColumnsRemain.remove(tableColumn)
                        return self.__linkByName(linkedColumns,notAssociate,similar)
            return linkedColumns

        def __linkByOrder(self,linkedColumns={},notAssociate={}):
            for column in self.__columnsRemain:
                if self.__pullColumnsRemain and not (column in notAssociate and notAssociate[column] == self.__pullColumnsRemain[0]):
                    linkedColumns[column] = self.__pullColumnsRemain[0]
                    self.__pullColumnsRemain.pop(0)
            return linkedColumns

        def __convertAssociations(self,table,associations={}):#Allow associations to be int, letter or name of Column in any order
            newAssociations = {}
            for association in associations:
                selfColumn = self.__matchColumn(association)#Search a column that matches with association in both tables
                tableColumn = table.__matchColumn(associations[association])

                if selfColumn == None or tableColumn == None:#If search not match both tables
                    selfColumn = self.__matchColumn(associations[association])#Reverse dict to test again
                    tableColumn = table.__matchColumn(association)
                
                if selfColumn != None and tableColumn != None:
                    newAssociations[selfColumn] = tableColumn
            return newAssociations
        
        def __matchColumn(self,testColumn):#Returns a column that matches with a testColumn (Associate by name, index or letter)
            for column in self.columns:
                if column.name == testColumn or column.index == testColumn or self.excel.convertColumn(column.index) == testColumn:
                    return column
            return None

        def __pullColumns(self,table,linkedColumns,replaceOldValues:bool,cut:bool):
            lastRow = self.lastRow
            for column in linkedColumns:
                selfRow = self.firstRow if replaceOldValues else (self.lastRow + 1 if self.firstRow!=self.lastRow else self.firstRow)
                tableRow = table.firstRow
                for cell in table.excel.getColumn(table.sheet,linkedColumns[column].index):
                    if cell.row == tableRow:
                        self.excel.setCellValue(self.sheet,column.index,selfRow,cell.value)
                        if cut:
                            table.excel.setCellValue(table.sheet,linkedColumns[column].index,tableRow,None)
                        selfRow += 1
                        tableRow += 1
                        lastRow = max(lastRow,selfRow)
            self.lastRow = max(lastRow,self.lastRow)

        def __getValuesFromMainColumn(self) -> list:
            values = []
            for cell in self.excel.getColumn(self.sheet,self.columns[self.mainColumn].index):
                if cell.row >= self.firstRow and cell.row <= self.lastRow:
                    values.append(cell.value)
            return values

        def __getMoreSimilar(self,value,allowedValues:list,autoCorrectWhen:float):
            if value in allowedValues:
                return value
            
            bestValue,bestValueSimilarity = None, 0
            for testValue in allowedValues:
                similarity = textSimilarity(simplifyText(str(value).strip()),simplifyText(str(testValue).strip()))
                if similarity > bestValueSimilarity:
                    bestValue,bestValueSimilarity = testValue,similarity
            if bestValueSimilarity > autoCorrectWhen:
                return bestValue
            return str(value) + ' (Not Found)'

    def getTable(self,sheet:str,headerRow:int,mainColumn=1,lastRow=-1) -> Table:#Returns a table of values
        return self.Table(self,sheet,headerRow,mainColumn,lastRow)

class Pdf:
    current_path = os.getcwd()
    pytesseract.pytesseract.tesseract_cmd = rf'{current_path}\lowModel\exterior\Tesseract-OCR\tesseract.exe'
    popplerBinPath = rf'{current_path}\lowModel\exterior\poppler-24.02.0\Library\bin'

    def merge(pdfs:list,outputPath):
        merger = PdfMerger()
        for pdf in pdfs:
            merger.append(pdf)
        merger.write(outputPath)
        merger.close()

    def readPdf(path) -> str:
        reader = PdfReader(path)
        text = ''
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text

    def readScannedPdf(path) -> str:
        pdfs = glob.glob(path)
        for pdf_path in pdfs:
            pages = convert_from_path(pdf_path, 500, poppler_path=Pdf.popplerBinPath)
            text = ''
            for imgBlob in pages:
                text += pytesseract.image_to_string(imgBlob,lang='por')
            return text
    
    def saveXML2003(pdfPath,xmlPath):
        word = win32com.client.gencache.EnsureDispatch("Word.Application")
        word.Visible = False
        word.Documents.Open(pdfPath)
        word.ActiveDocument.SaveAs(xmlPath,11)
        word.Quit()

