from lowModel.utils import *
import openpyxl as xl
import win32com.client
from math import ceil
from datetime import date
from shutil import copyfile

from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from pypdf import PdfMerger
from pypdf import PdfReader
from pdf2image import convert_from_path
import pytesseract
import glob

class Excel:
    def __init__(self,pathToXlsx:str):
        self.wb = xl.load_workbook(pathToXlsx)#OpenPyXl WorkBook
        self.path = pathToXlsx
        self.fileName = self.getFileNameByFilePath(pathToXlsx)
        self.folderPath = self.getFolderPathByFilePath(pathToXlsx)
        self.savesBackupFolder = self.getBackupFolderByPath(pathToXlsx)

    def getFileNameByFilePath(self,path:str):#Returns the text between the last slash and the .xlsx (FileName)
        return path[path.rfind('\\')+1:path.rfind('.')]
        
    def getFolderPathByFilePath(self,path:str):#Returns all text before the last slash (FolderPath)
        return path[:path.rfind('\\')+1]
    
    def getBackupFolderByPath(self,path:str):#Returns the path to a new folder where backups of the modified Excel will be stored
        return self.getFolderPathByFilePath(path) + '0oldVersions\\'

    def getSheets(self):#Returns the name of WorkSheets
        return self.wb.sheetnames
    
    def getRow(self,sheet:str,row:int):#Returns a list of Cells in Row (Begins with 1)
        return list(self.wb[sheet].rows[row-1])
    
    def getColumn(self,sheet:str,column:int):#Returns a list of Cells in Column (Begins with 1)
        return list(self.wb[sheet].columns)[column-1]
    
    def getRowOfValue(self,sheet:str,column:int,value,occurrenceIndex=1):#Return the Row of Value in Column (Index -1 returns all occurrences)
        occurrences = []
        for row,cell in enumerate(self.getColumn(sheet,column)):
            if cell.value == value:
                occurrences.append(row)
                if len(occurrences) == occurrenceIndex:
                    return row
        if occurrenceIndex == -1:
            return occurrences
        return None
    
    def getColumnOfValue(self,sheet:str,row:int,value,occurrenceIndex=1):#Return the Column of Value in Row (Index -1 returns all occurrences)
        occurrences = []
        for column,cell in enumerate(self.getRow(sheet,row)):
            if cell.value == value:
                occurrences.append(column)
                if len(occurrences) == occurrenceIndex:
                    return column
        if occurrenceIndex == -1:
            return occurrences
        return None
    
    def getCellValue(self,sheet:str,row:int,column,allowFormula=True):#Returns Cell Value (Row begins with 1 and Column can be int or str)
        column = self.convertColumn(column)
        if allowFormula:
            return self.wb[sheet].cell(row,column).value
        else:
            readWb = xl.load_workbook(self.path,read_only=True,data_only=True)
            result = readWb[sheet].cell(row,column).value
            readWb.close()
            return result
        
    def setCellValue(self,sheet:str,row:int,column,value):#Change Cell Value (needs "save()")
        column = self.convertColumn(column)
        self.wb[sheet].cell(row,column).value = value
        '''
            cellFormat = self.wb[sheet][cell].number_format
            self.wb[sheet][cell].value = value
            self.wb[sheet][cell].number_format = cellFormat
        '''

    def setHide(self,sheet:str,row=None,column=None,hide=True):#Hide or Unhide a row/column
        if row:
            self.wb[sheet].row_dimensions[row].hidden = hide
        if column:
            column = self.convertColumn(column)
            self.wb[sheet].column_dimensions[column].hidden = hide

    def convertColumn(self,column:str) -> int:#If column is str convert to int ('A'=1,'B'=2)
        if type(column) == str:
            pass
        return column

class OldExcel:
    def __init__(self,path):
        self.path = path
        self.fileName = path[path.rfind('\\')+1:path.rfind('.')]
        self.fileFolder = path.replace(f'{self.fileName}.xlsx','')
        self.backupFolder = self.fileFolder+'0oldVersions\\'
        self.wb = xl.load_workbook(path)
    
    def sheets(self):
        return self.wb.worksheets

    def getRow(self,sheet,row):
        return list(self.wb[sheet].rows)[row]
    
    def getColumn(self,sheet,column):
        return list(self.wb[sheet].columns)[column]

    def getRowOfValue(self,sheet,column,valueSearch):
        for n,cell in enumerate(list(self.wb[sheet].columns)[column]):
            if cell.value == valueSearch:
                return n
        return None
    
    def getColumnOfValue(self,sheet,row,valueSearch):
        for n,cell in enumerate(list(self.wb[sheet].rows)[row]):
            if cell.value == valueSearch:
                return n
        return None

    def get(self,sheet,cell,formula=False):
        if formula:
            return self.wb[sheet][cell].value
        else:
            read = xl.load_workbook(self.path,data_only=True,read_only=True)
            result = read[sheet][cell].value
            read.close()
            return result
        
    def getByIndex(self,sheet,row,column,formula=False):
        if formula:
            return self.wb[sheet].cell(row+1,column+1).value
        else:
            read = xl.load_workbook(self.path,data_only=True,read_only=True)
            result = read[sheet].cell(row+1,column+1).value
            read.close()
            return result

    def set(self,sheet,cell,value,resize=True):
        cellFormat = self.wb[sheet][cell].number_format
        self.wb[sheet][cell].value = value
        self.wb[sheet][cell].number_format = cellFormat
        if resize:
            self.resize(sheet,cell,value)

    def setByIndex(self,sheet,row,column,value,resize=True):
        self.wb[sheet].cell(row+1,column+1).value = value
        if resize:
            self.resizeIndex(sheet,row,column,value)

    def resize(self,sheet,cell,value,maxCharInCell=77,lineSep=12,minHeight=36,linesNotFormat=3):
        row = self.cellToRow(cell)
        self.wb[sheet].row_dimensions[row].height = minHeight
        linesToAdd = ceil(len(str(value))/maxCharInCell) - linesNotFormat
        self.wb[sheet].row_dimensions[row].height += max(0,lineSep * linesToAdd)

    def resizeIndex(self,sheet,row,column,value,maxCharInCell=77,lineSep=12,minHeight=36,linesNotFormat=3):
        self.wb[sheet].row_dimensions[row+1].height = minHeight
        linesToAdd = ceil(len(str(value))/maxCharInCell) - linesNotFormat
        self.wb[sheet].row_dimensions[row+1].height += max(0,lineSep * linesToAdd)

    def setHide(self,sheet,cell,value:bool):
        cell = self.cellToRow(cell)
        self.wb[sheet].row_dimensions[cell].hidden = value

    def setHideIndex(self,sheet,row,column,value:bool):
        self.wb[sheet].row_dimensions[row+1].hidden = value

    def formatByIndex(self,sheet,row,column,name=None,color=None,size=None):
        cell = self.wb[sheet].cell(row+1,column+1)
        if color:
            cell.font = Font(name=name,color=color,size=size)

    def save(self,backup=True,reopen=True):
        if backup:
            self.backup()
        self.wb.save(self.path)
        if reopen:
            self.wb = xl.load_workbook(self.path)

    def savePdf(self,pages,path=None,backup=True):
        if path == None:
            path = self.path.replace('xlsx','pdf')
        self.save(backup)
        pages = self.convertSheetRange(pages)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(self.path)
        ws_index_list = pages
        wb.WorkSheets(ws_index_list).Select()
        wb.ActiveSheet.ExportAsFixedFormat(0,path)
        os.system("taskkill /f /im excel.exe")

    def backup(self):
        today = date.today().strftime('%d.%m')

        if not os.path.exists(self.backupFolder):
            os.mkdir(self.backupFolder)
        
        version = ''
        numBackups = 0
        while True:
            filePath = f'{self.backupFolder}{self.fileName} {today}{version}.xlsx'
            if os.path.exists(filePath):
                numBackups += 1
                version = f' ({numBackups})'
            else:
                return copyfile(self.path,filePath)
            
    def cellToRow(self,cell):
        if type(cell) == str:
            value = 0
            for c in cell:
                if c.isnumeric():
                    value = 10*value + int(c)
            return value
        return cell
    
    def convertSheetRange(self,sheetRange):
        if sheetRange == -1:
            return range(1,len(self.wb.worksheets)+1)
        elif type(sheetRange) == str:
            return [self.sheetIndex(sheetRange)+1]
        elif type(sheetRange) != list:
            return [sheetRange]
        return sheetRange
    
    def sheetIndex(self,sheetName):
        for n,sheet in enumerate(self.wb.worksheets):
            if sheet.title == sheetName:
                return n
            
    def close(self):
        self.wb.close()

class OldPdf:
    current_path = os.getcwd()
    pytesseract.pytesseract.tesseract_cmd = rf'{current_path}\lowModel\exterior\Tesseract-OCR\tesseract.exe'
    popplerBinPath = rf'{current_path}\lowModel\exterior\poppler-24.02.0\Library\bin'

    def merge(pdfs:list,outputPath):
        merger = PdfMerger()
        for pdf in pdfs:
            merger.append(pdf)
        merger.write(outputPath)
        merger.close()

    def readPdf(path) -> str:
        reader = PdfReader(path)
        text = ''
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text

    def readScannedPdf(path) -> str:
        pdfs = glob.glob(path)
        for pdf_path in pdfs:
            pages = convert_from_path(pdf_path, 500, poppler_path=OldPdf.popplerBinPath)
            text = ''
            for imgBlob in pages:
                text += pytesseract.image_to_string(imgBlob,lang='por')
            return text
    
    def saveXML2003(pdfPath,xmlPath):
        word = win32com.client.gencache.EnsureDispatch("Word.Application")
        word.Visible = False
        word.Documents.Open(pdfPath)
        word.ActiveDocument.SaveAs(xmlPath,11)
        word.Quit()

